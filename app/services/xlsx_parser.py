import io
import openpyxl
from app.validators.xlsx_validator import (
    FracaoRow,
    CabecalhoRow,
    validate_fracoes,
    validate_cabecalho,
)


def parse_xlsx(file_bytes: bytes) -> tuple[list[FracaoRow], list[CabecalhoRow]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    fracoes_raw = _sheet_to_dicts(wb, "fracoes")
    cabecalho_raw = _sheet_to_dicts(wb, "cabecalho")

    wb.close()

    fracoes = validate_fracoes(fracoes_raw)
    cabecalho = validate_cabecalho(cabecalho_raw)

    return fracoes, cabecalho


def _sheet_to_dicts(wb: openpyxl.Workbook, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        if sheet_name == "fracoes":
            raise ValueError(f"Aba '{sheet_name}' nao encontrada no arquivo")
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    result: list[dict] = []

    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        record = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            value = row[col_idx] if col_idx < len(row) else None
            record[header] = value if value is not None else ""
        result.append(record)

    return result
